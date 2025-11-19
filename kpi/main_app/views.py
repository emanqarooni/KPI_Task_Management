from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse
from .models import (
    EmployeeKpi,
    ProgressEntry,
    Kpi,
    EmployeeProfile,
    DEPARTMENT,
    ActivityLog,
    Notification,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import CreateView, UpdateView, View
from .forms import AssignKpiForm, KpiProgressForm
from django.contrib import messages
from .decorators import RoleRequiredMixin, role_required
from django.contrib.auth.models import User
from django.db.models import Sum, Q, Count
from .services.ai import generate_kpi_insights
import markdown
from django.utils.timezone import now
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
from django.contrib.admin.views.decorators import staff_member_required
from decimal import Decimal


import google.generativeai as genai

# create your views here


# Employee dashboard'
@login_required
def dashboard(request):
    employee_profile = EmployeeProfile.objects.get(user=request.user)
    if employee_profile.role == "admin":
        return redirect("admin_dashboard")
    elif employee_profile.role == "manager":
        return redirect("manager_dashboard")
    else:
        return redirect("employee_dashboard")


# Admin dashboard
@login_required
@role_required(["admin"])
def admin_dashboard(request):
    # Total users
    total_users_count = User.objects.count()

    # Employees & managers
    all_employees = EmployeeProfile.objects.filter(role="employee").select_related(
        "user"
    )
    total_employees_count = all_employees.count()

    all_managers = EmployeeProfile.objects.filter(role="manager").select_related("user")
    total_managers_count = all_managers.count()

    # Department counts
    department_counts = EmployeeProfile.objects.values("department").annotate(
        count=Count("id")
    )

    department_data = {
        "Sales_Marketing": 0,
        "Information_Technology": 0,
        "Human_Resources": 0,
        "Project_Management": 0,
    }
    department_map = {
        "SM": "Sales_Marketing",
        "IT": "Information_Technology",
        "HR": "Human_Resources",
        "PM": "Project_Management",
    }
    for dept in department_counts:
        dept_code = dept["department"]
        key = department_map.get(dept_code, dept_code)
        department_data[key] = dept["count"]

    all_kpis = Kpi.objects.all()
    employee_kpi_assignments = EmployeeKpi.objects.select_related(
        "employee__user", "employee__manager", "kpi"
    )
    employee_kpi_data = []
    for assignment in employee_kpi_assignments:
        progress_percentage = assignment.progress_percentage()
        weighted_score = (progress_percentage * float(assignment.weight)) / 100
        employee_kpi_data.append(
            {
                "employee_name": assignment.employee.user.get_full_name()
                or assignment.employee.user.username,
                "manager_name": (
                    assignment.employee.manager.get_full_name()
                    if assignment.employee.manager
                    else "No Manager"
                ),
                "kpi_title": assignment.kpi.title,
                "target_value": assignment.target_value,
                "current_progress": assignment.total_progress(),
                "progress_percentage": progress_percentage,
                "weight": assignment.weight,
                "status": assignment.status(),
                "weighted_score": round(weighted_score, 2),
            }
        )

    employee_total_scores = {}
    for data in employee_kpi_data:
        name = data["employee_name"]
        employee_total_scores[name] = (
            employee_total_scores.get(name, 0) + data["weighted_score"]
        )

    chart_labels_list = []
    chart_values_list = []

    for kpi in all_kpis:
        chart_labels_list.append(kpi.title)
        total_progress_for_kpi = (
            ProgressEntry.objects.filter(employee_kpi__kpi=kpi).aggregate(
                total=Sum("value")
            )["total"]
            or 0
        )
        chart_values_list.append(float(total_progress_for_kpi))
    context = {
        "total_users": total_users_count,
        "total_employees": total_employees_count,
        "total_managers": total_managers_count,
        "department_data": department_data,
        "employees": all_employees,
        "managers": all_managers,
        "employee_kpi_data": employee_kpi_data,
        "employee_total_scores": employee_total_scores,  # <- added
        "chart_labels": json.dumps(chart_labels_list),
        "chart_values": json.dumps([float(v) for v in chart_values_list]),
    }
    return render(request, "dashboards/admin_dashboard.html", context)


# Manager dashboard
@login_required
@role_required(["manager"])
def manager_dashboard(request):
    manager_profile = EmployeeProfile.objects.get(user=request.user)
    manager_department = manager_profile.department
    employees_in_department = EmployeeProfile.objects.filter(
        department=manager_department, role="employee"
    )
    employee_dashboard_rows = []
    for employee in employees_in_department:
        employee_kpis = EmployeeKpi.objects.filter(employee=employee)
        if employee_kpis.exists():
            total_progress_value = 0
            total_target_value = 0
            for employee_kpi in employee_kpis:
                progress_entries = employee_kpi.progressentry_set.all()
                for entry in progress_entries:
                    total_progress_value += entry.value
                total_target_value += employee_kpi.target_value
            if total_target_value == 0:
                completion_percentage = 0
            else:
                completion_percentage = round(
                    (total_progress_value / total_target_value) * 100
                )
            if completion_percentage == 0:
                status_text = "No Progress"
            elif completion_percentage >= 100:
                status_text = "Complete"
            else:
                status_text = "Moderate"
        else:
            completion_percentage = 0
            status_text = "No KPI Assigned"
        employee_dashboard_rows.append(
            {
                "name": employee.user.username,
                "completion": completion_percentage,
                "status": status_text,
            }
        )
    context = {
        "manager_name": manager_profile.user.username,
        "department": manager_profile.get_department_display(),
        "employees_count": employees_in_department.count(),
        "employee_rows": employee_dashboard_rows,
    }
    return render(request, "dashboards/manager_dashboard.html", context)


# Employee dashboard
@login_required
@role_required(["employee"])
def employee_dashboard(request):
    employee_profile = EmployeeProfile.objects.get(user=request.user)
    assigned_kpis = EmployeeKpi.objects.filter(employee=employee_profile)
    chart_labels_list = []
    chart_values_list = []
    kpi_cards_list = []
    for employee_kpi in assigned_kpis:
        kpi_title = employee_kpi.kpi.title
        total_progress_value = 0
        progress_entries = employee_kpi.progressentry_set.all()
        for entry in progress_entries:
            total_progress_value += entry.value
        if employee_kpi.target_value == 0:
            progress_percentage = 0
        else:
            progress_percentage = round(
                (total_progress_value / employee_kpi.target_value) * 100
            )
        days_remaining = (employee_kpi.end_date - date.today()).days
        chart_labels_list.append(kpi_title)
        chart_values_list.append(progress_percentage)
        kpi_cards_list.append(
            {
                "title": kpi_title,
                "target": employee_kpi.target_value,
                "progress": total_progress_value,
                "percentage": progress_percentage,
                "days_left": days_remaining,
                "id": employee_kpi.id,
            }
        )
    context = {
        "employee_name": employee_profile.user.username,
        "role": employee_profile.job_role,
        "kpi_cards": kpi_cards_list,
        "chart_labels": json.dumps(chart_labels_list),
        "chart_values": json.dumps(chart_values_list),
    }
    return render(request, "dashboards/employee_dashboard.html", context)


def home(request):
    return render(request, "home.html")


@login_required
@role_required(["admin"])
def kpis_index(request):
    kpis = Kpi.objects.all()
    return render(request, "kpi/index.html", {"kpis": kpis})


@login_required
@role_required(["admin"])
def kpis_detail(request, kpi_id):
    kpi = Kpi.objects.get(id=kpi_id)
    return render(request, "kpi/detail.html", {"kpi": kpi})


class KpiCreate(RoleRequiredMixin, CreateView):
    model = Kpi
    fields = "__all__"
    success_url = "/kpis/"
    allowed_roles = ["admin"]


@login_required
@role_required(["employee"])
def add_progress(request):
    if request.method == "POST":
        form = KpiProgressForm(request.POST, user=request.user)

        if form.is_valid():
            employee_kpi = form.cleaned_data["employee_kpi"]

            if employee_kpi.status() == "Complete":
                messages.error(
                    request,
                    "This KPI is already complete. You cannot add more progress.",
                )
                return redirect("employee_kpi")

            progress = form.save()

            # check if the kpi is complete
            is_completed = employee_kpi.status() == "Complete"

            # send notification to manager
            if employee_kpi.employee.manager:
                notification_title = (
                    "KPI Completed!" if is_completed else "Progress Updated"
                )
                notification_message = (
                    f"{request.user.get_full_name() or request.user.username} has completed the KPI: {employee_kpi.kpi.title}"
                    if is_completed
                    else f'{request.user.get_full_name() or request.user.username} added progress to "{employee_kpi.kpi.title}". '
                    f"Current progress: {employee_kpi.progress_percentage()}%"
                )

                create_notification(
                    recipient=employee_kpi.employee.manager,
                    sender=request.user,
                    notification_type=(
                        "kpi_completed" if is_completed else "progress_added"
                    ),
                    title=notification_title,
                    message=notification_message,
                    employee_kpi=employee_kpi,
                )

            ActivityLog.create_log(
                user=request.user,
                action="PROGRESS_ADDED",
                description=f"Added progress for '{progress.employee_kpi.kpi.title}' - Value: {progress.value}",
                related_user=None,
            )
            messages.success(request, "Progress added successfully!")
            return redirect("employee_kpi")

    else:
        form = KpiProgressForm(user=request.user)

    return render(request, "kpi/progress.html", {"form": form})


@login_required
@role_required(["employee"])
def employee_kpi(request):
    employee_kpi = EmployeeKpi.objects.all()
    return render(request, "kpi/employee_kpi.html", {"employee_kpi": employee_kpi})


class KpiUpdate(RoleRequiredMixin, UpdateView):
    model = Kpi
    fields = "__all__"
    allowed_roles = ["admin"]


# unauthorized access page
def unauthorized(request):
    return render(request, "unauthorized.html")


@login_required
def profile(request):
    return render(request, "users/profile.html")


# assigning kpi
@login_required
@role_required(["manager"])
def assign_kpi(request):
    # pass request.user into the form so it filters by department if user is manager
    form = AssignKpiForm(
        request.POST or None,
        user=(request.user if request.user.is_authenticated else None),
    )
    if request.method == "POST":
        if form.is_valid():
            kpi_assignment = form.save()

            ActivityLog.create_log(
                user=request.user,
                action="KPI_ASSIGNED",
                description=f"Assigned '{kpi_assignment.kpi.title}' to {kpi_assignment.employee.user.username} (Target: {kpi_assignment.target_value}, Weight: {kpi_assignment.weight}%)",
                related_user=kpi_assignment.employee.user,
            )

            # send notification to employees
            create_notification(
                recipient=kpi_assignment.employee.user,
                sender=request.user,
                notification_type="kpi_assigned",
                title="New KPI Assigned",
                message=f"You have been assigned a new KPI: {kpi_assignment.kpi.title}. Target: {kpi_assignment.target_value}",
                employee_kpi=kpi_assignment,
            )

            messages.success(request, "KPI assigned successfully.")
            return redirect("employee_kpi_list")
    return render(request, "main_app/assign_kpi.html", {"form": form})


@login_required
@role_required(["manager"])
def employee_kpi_list(request):
    if (
        request.user.is_authenticated
        and hasattr(request.user, "employeeprofile")
        and request.user.employeeprofile.role == "manager"
    ):
        dept = request.user.employeeprofile.department
        # show only assignments where employee is in manager's department
        kpis = (
            EmployeeKpi.objects.filter(employee__department=dept)
            .select_related("employee__user", "kpi")
            .order_by("-id")
        )
    else:
        kpis = EmployeeKpi.objects.select_related("employee__user", "kpi").all()

    # search functionality
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
            | Q(employee__job_role__icontains=search_query)
        )

    # filter by KPI
    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    # order by newest first
    kpis = kpis.order_by("-id")

    status_filter = request.GET.get("status", "")
    if status_filter:
        # convert the queryset to a list and filter by status since status is a method
        filtered_kpis = []
        for kpi in kpis:
            kpi_status = kpi.status().lower().replace(" ", "_")
            if kpi_status == status_filter:
                filtered_kpis.append(kpi)
        kpis = filtered_kpis
    else:
        # convert to list for consistency
        kpis = list(kpis)

    # get all KPIs for the filter dropdown
    if (
        request.user.is_authenticated
        and hasattr(request.user, "employeeprofile")
        and request.user.employeeprofile.role == "manager"
    ):
        dept = request.user.employeeprofile.department
        all_kpis = Kpi.objects.filter(department=dept)
    else:
        all_kpis = Kpi.objects.all()

    context = {
        "kpis": kpis,
        "all_kpis": all_kpis,
        "search_query": search_query,
        "kpi_filter": kpi_filter,
        "status_filter": status_filter,
    }
    return render(request, "main_app/employee_kpi_list.html", context)


@login_required
@role_required(["manager"])
def employee_kpi_edit(request, pk):
    kpi_assign = get_object_or_404(EmployeeKpi, pk=pk)
    if kpi_assign.progressentry_set.exists():
        messages.error(request, "Cannot edit KPI — progress already exists.")
        return redirect("employee_kpi_list")

    form = AssignKpiForm(
        request.POST or None,
        instance=kpi_assign,
        user=(request.user if request.user.is_authenticated else None),
    )
    if request.method == "POST" and form.is_valid():
        kpi_assignment = form.save()

        ActivityLog.create_log(
            user=request.user,
            action="KPI_UPDATED",
            description=f"Updated KPI '{kpi_assignment.kpi.title}' for {kpi_assignment.employee.user.username}",
            related_user=kpi_assignment.employee.user,
        )

        # send notification to employee
        create_notification(
            recipient=employee_kpi.employee.user,
            sender=request.user,
            notification_type="kpi_updated",
            title="KPI Updated",
            message=f'Your KPI "{employee_kpi.kpi.title}" has been updated by your manager.',
            employee_kpi=employee_kpi,
        )

        messages.success(request, "KPI updated successfully.")
        return redirect("employee_kpi_list")

    return render(
        request, "main_app/employee_kpi_edit.html", {"form": form, "kpi": kpi_assign}
    )


@login_required
@role_required(["manager"])
def employee_kpi_delete(request, pk):
    kpi_assign = get_object_or_404(EmployeeKpi, pk=pk)
    if kpi_assign.progressentry_set.exists():
        messages.error(request, "Cannot delete KPI — progress already exists.")
        return redirect("employee_kpi_list")

    if request.method == "POST":
        employee_name = kpi_assign.employee.user.username
        kpi_title = kpi_assign.kpi.title
        related_user = kpi_assign.employee.user

        ActivityLog.create_log(
            user=request.user,
            action="KPI_DELETED",
            description=f"Deleted KPI assignment '{kpi_title}' from {employee_name}",
            related_user=related_user,
        )

        kpi_assign.delete()
        messages.success(request, "KPI assignment deleted.")
        return redirect("employee_kpi_list")

    return render(request, "main_app/employee_kpi_delete.html", {"kpi": kpi_assign})


@login_required
def employee_kpi_detail(request, pk):
    profile = request.user.employeeprofile

    # managers can view any kpi in their department
    if profile.role == "manager":
        kpi_assign = get_object_or_404(
            EmployeeKpi, pk=pk, employee__department=profile.department
        )
    # employees can only view their own KPIs
    elif profile.role == "employee":
        kpi_assign = get_object_or_404(EmployeeKpi, pk=pk, employee__user=request.user)
    else:
        return redirect("unauthorized")

    progress_entries = kpi_assign.progressentry_set.order_by("date")
    return render(
        request,
        "main_app/employee_kpi_detail.html",
        {
            "kpi": kpi_assign,
            "progress_entries": progress_entries,
        },
    )


# reports views


# manager reports view with filtering
@login_required
@role_required(["manager"])
def manager_reports(request):
    dept = request.user.employeeprofile.department

    # get kpis that r filtered by manager's department
    kpis = (
        EmployeeKpi.objects.filter(employee__department=dept)
        .select_related("employee__user", "kpi")
        .order_by("-id")
    )

    # apply filters from get parameters
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
        )

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # get all kpis for the filter dropdown
    all_kpis = Kpi.objects.filter(department=dept)

    context = {
        "kpis": kpis,
        "all_kpis": all_kpis,
        "search_query": search_query,
        "kpi_filter": kpi_filter,
        "start_date": start_date,
        "end_date": end_date,
    }
    return render(request, "reports/manager_reports.html", context)


# admin reports view with filtering across all departments
@login_required
@role_required(["admin"])
def admin_reports(request):
    # get all kpis across all departments
    kpis = EmployeeKpi.objects.select_related(
        "employee__user",
        "kpi",
        "employee",
        "employee__manager",
        "employee__manager__employeeprofile",
    ).order_by("-id")

    # apply filters from get parameters
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
            | Q(employee__manager__username__icontains=search_query)
            | Q(employee__manager__first_name__icontains=search_query)
            | Q(employee__manager__last_name__icontains=search_query)
        )

    # dept filterations
    department_filter = request.GET.get("department", "")
    if department_filter:
        kpis = kpis.filter(employee__department=department_filter)

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # get all kpis for the filter dropdown (all departments)
    all_kpis = Kpi.objects.all()

    # get all departments for filter
    all_departments = DEPARTMENT
    context = {
        "kpis": kpis,
        "all_kpis": all_kpis,
        "all_departments": all_departments,
        "search_query": search_query,
        "kpi_filter": kpi_filter,
        "department_filter": department_filter,
        "start_date": start_date,
        "end_date": end_date,
    }
    return render(request, "reports/admin_reports.html", context)


# exporting pdfs from manager side
@login_required
@role_required(["manager"])
def export_pdf(request):
    dept = request.user.employeeprofile.department

    # get filtered kpis same logic as manager_reports
    kpis = (
        EmployeeKpi.objects.filter(employee__department=dept)
        .select_related("employee__user", "kpi")
        .order_by("-id")
    )

    # apply filters
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
        )

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # create PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename=KPI_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

    # create PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []

    # Styles
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        f"<b>KPI Performance Report</b><br/>{datetime.now().strftime('%B %d, %Y')}",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    # table data
    data = [["KPI", "Employee", "Target", "Current", "Progress", "Status", "Period"]]

    for kpi in kpis:
        data.append(
            [
                kpi.kpi.title,
                kpi.employee.user.first_name + " " + kpi.employee.user.last_name,
                str(kpi.target_value),
                str(kpi.total_progress()),
                f"{kpi.progress_percentage()}%",
                kpi.status(),
                f"{kpi.start_date}\n - {kpi.end_date}",
            ]
        )

    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                # Body styling
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                # Alternating rows
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


# exporting excel from manager side
@login_required
@role_required(["manager"])
def export_excel(request):
    dept = request.user.employeeprofile.department

    # get filtered kpis same logic as manager_reports
    kpis = (
        EmployeeKpi.objects.filter(employee__department=dept)
        .select_related("employee__user", "kpi")
        .order_by("-id")
    )

    # apply filters from get parameter
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
        )

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # createing workbook
    wb = Workbook()

    # select the active sheet
    sheet = wb.active
    sheet.title = "KPI Report"

    # header styling
    header_fill = PatternFill(
        start_color="0F172A", end_color="0F172A", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # title
    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = f"KPI Performance Report - {datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # headers
    headers = [
        "KPI",
        "Employee",
        "Target",
        "Current Value",
        "Progress %",
        "Status",
        "Period",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # data
    for row_num, kpi in enumerate(kpis, 4):
        sheet.cell(row=row_num, column=1).value = kpi.kpi.title
        sheet.cell(row=row_num, column=2).value = (
            kpi.employee.user.first_name + " " + kpi.employee.user.last_name
        )
        sheet.cell(row=row_num, column=3).value = kpi.target_value
        sheet.cell(row=row_num, column=4).value = kpi.total_progress()
        sheet.cell(row=row_num, column=5).value = f"{kpi.progress_percentage()}%"
        sheet.cell(row=row_num, column=6).value = kpi.status()
        sheet.cell(row=row_num, column=7).value = f"{kpi.start_date} → {kpi.end_date}"

    # auto-adjust column widths
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = None

        for cell in column_cells:
            # skip merged cells
            if hasattr(cell, "column_letter"):
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

        if column_letter and max_length > 0:
            adjusted_width = min(
                max_length + 2, 50
            )  # here capping at 50 to avoid overly wide columns
            sheet.column_dimensions[column_letter].width = adjusted_width

    # prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=KPI_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response


# Admin export to PDF
@login_required
@role_required(["admin"])
def admin_export_pdf(request):
    # get filtered kpis - same logic as admin_reports
    kpis = EmployeeKpi.objects.select_related(
        "employee__user", "kpi", "employee__manager"
    ).order_by("-id")

    # apply filters
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
            | Q(employee__manager__username__icontains=search_query)
            | Q(employee__manager__first_name__icontains=search_query)
            | Q(employee__manager__last_name__icontains=search_query)
        )

    department_filter = request.GET.get("department", "")
    if department_filter:
        kpis = kpis.filter(employee__department=department_filter)

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # create PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename=Admin_KPI_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

    # create PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []

    # Styles
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        f"<b>Admin KPI Performance Report</b><br/>{datetime.now().strftime('%B %d, %Y')}",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    # table data
    data = [
        [
            "KPI",
            "Employee",
            "Manager",
            "Dept",
            "Target",
            "Current",
            "Progress",
            "Status",
            "Period",
        ]
    ]

    for kpi in kpis:
        manager_name = (
            f"{kpi.employee.manager.first_name} {kpi.employee.manager.last_name}"
            if kpi.employee.manager
            else "N/A"
        )
        employee_name = f"{kpi.employee.user.first_name} {kpi.employee.user.last_name}"
        data.append(
            [
                kpi.kpi.title[:20],
                employee_name,
                manager_name,
                kpi.employee.get_department_display()[:35],
                str(kpi.target_value),
                str(kpi.total_progress()),
                f"{kpi.progress_percentage()}%",
                kpi.status(),
                f"{kpi.start_date}\n{kpi.end_date}",
            ]
        )

    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                # Body styling
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                # Alternating rows
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]
        )
    )

    elements.append(table)

    # Build PDF
    doc.build(elements)

    return response


# Admin export to Excel
@login_required
@role_required(["admin"])
def admin_export_excel(request):
    # get filtered kpis - same logic as admin_reports
    kpis = EmployeeKpi.objects.select_related(
        "employee__user", "kpi", "employee__manager"
    ).order_by("-id")

    # apply filters from get parameter
    search_query = request.GET.get("search", "").strip()
    if search_query:
        kpis = kpis.filter(
            Q(employee__user__username__icontains=search_query)
            | Q(employee__user__first_name__icontains=search_query)
            | Q(employee__user__last_name__icontains=search_query)
            | Q(employee__manager__username__icontains=search_query)
            | Q(employee__manager__first_name__icontains=search_query)
            | Q(employee__manager__last_name__icontains=search_query)
        )

    department_filter = request.GET.get("department", "")
    if department_filter:
        kpis = kpis.filter(employee__department=department_filter)

    kpi_filter = request.GET.get("kpi", "")
    if kpi_filter:
        kpis = kpis.filter(kpi__id=kpi_filter)

    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if start_date:
        kpis = kpis.filter(start_date__gte=start_date)
    if end_date:
        kpis = kpis.filter(end_date__lte=end_date)

    # creating workbook
    wb = Workbook()

    # select the active sheet
    sheet = wb.active
    sheet.title = "Admin KPI Report"

    # header styling
    header_fill = PatternFill(
        start_color="0F172A", end_color="0F172A", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # title
    sheet.merge_cells("A1:J1")
    title_cell = sheet["A1"]
    title_cell.value = (
        f"Admin KPI Performance Report - {datetime.now().strftime('%Y-%m-%d')}"
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # headers
    headers = [
        "KPI",
        "Employee",
        "Manager",
        "Department",
        "Target",
        "Current Value",
        "Progress %",
        "Weight %",
        "Status",
        "Period",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # data
    for row_num, kpi in enumerate(kpis, 4):
        manager_name = (
            f"{kpi.employee.manager.first_name} {kpi.employee.manager.last_name}"
            if kpi.employee.manager
            else "No Manager"
        )
        employee_name = f"{kpi.employee.user.first_name} {kpi.employee.user.last_name}"

        sheet.cell(row=row_num, column=1).value = kpi.kpi.title
        sheet.cell(row=row_num, column=2).value = employee_name
        sheet.cell(row=row_num, column=3).value = manager_name
        sheet.cell(row=row_num, column=4).value = kpi.employee.get_department_display()
        sheet.cell(row=row_num, column=5).value = kpi.target_value
        sheet.cell(row=row_num, column=6).value = kpi.total_progress()
        sheet.cell(row=row_num, column=7).value = f"{kpi.progress_percentage()}%"
        sheet.cell(row=row_num, column=8).value = f"{kpi.weight}%"
        sheet.cell(row=row_num, column=9).value = kpi.status()
        sheet.cell(row=row_num, column=10).value = f"{kpi.start_date} → {kpi.end_date}"

    # auto-adjust column widths
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = None

        for cell in column_cells:
            if hasattr(cell, "column_letter"):
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

        if column_letter and max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=Admin_KPI_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response


@login_required
@role_required(["admin"])
def activity_logs(request):
    # get url values form filter form
    selected_action = request.GET.get("action", "")
    selected_user = request.GET.get("user", "")

    all_logs = ActivityLog.objects.all()
    all_logs = all_logs.select_related("user", "related_user")

    if selected_action:
        all_logs = all_logs.filter(action=selected_action)

    if selected_user:
        all_logs = all_logs.filter(
            Q(user__username__icontains=selected_user)
            | Q(user__first_name__icontains=selected_user)
            | Q(user__last_name__icontains=selected_user)
        )

    recent_logs = all_logs[:50]

    action_choices = ActivityLog.ACTION_CHOICES

    users_with_logs = User.objects.filter(activity_logs__isnull=False)
    users_with_logs = users_with_logs.select_related("employeeprofile")
    users_with_logs = users_with_logs.distinct()
    users_with_logs = users_with_logs.order_by("username")

    template_data = {
        "logs": recent_logs,
        "available_actions": action_choices,
        "active_users": users_with_logs,
        "current_action_filter": selected_action,
        "current_user_filter": selected_user,
    }
    return render(request, "activity/admin_logs.html", template_data)


#  this is the helper function to create notifications
def create_notification(
    recipient, sender, notification_type, title, message, employee_kpi=None
):
    Notification.objects.create(
        recipient=recipient,
        sender=sender,
        notification_type=notification_type,
        title=title,
        message=message,
        employee_kpi=employee_kpi,
    )


# view all notifications for current user
@login_required
def notifications(request):
    profile = EmployeeProfile.objects.get(user=request.user)

    # exclude admin from notifications
    if profile.role == "admin":
        return redirect("dashboard")

    user_notifications = Notification.objects.filter(recipient=request.user)
    unread_count = user_notifications.filter(is_read=False).count()

    context = {
        "notifications": user_notifications,
        "unread_count": unread_count,
    }
    return render(request, "notifications/notifications.html", context)


@login_required
def mark_notification_read(request, notification_id):
    # marking a single notification as read
    notification = get_object_or_404(
        Notification, id=notification_id, recipient=request.user
    )
    notification.is_read = True
    notification.save()
    return redirect("notifications")


@login_required
def mark_all_notifications_read(request):
    # marking all notifications as read
    Notification.objects.filter(recipient=request.user, is_read=False).update(
        is_read=True
    )
    messages.success(request, "All notifications marked as read.")
    return redirect("notifications")


## AI feature for Manager View
@login_required
@role_required(["manager"])
def ai_kpi_insights(request):
    manager_user = request.user

    employees = EmployeeProfile.objects.filter(
        manager=manager_user,
    )
    data = ""

    for emp in employees:
        data += f"\n=== Employee: {emp.user.get_full_name() or emp.user.username} ===\n"

        emp_kpis = EmployeeKpi.objects.filter(employee=emp)

        active_kpis = []
        previous_kpis = []

        for ek in emp_kpis:
            if ek.status() == "Complete":
                previous_kpis.append(ek)
            else:
                active_kpis.append(ek)

        # current active KPI (latest valid progress entry)
        current_kpi = None
        current_last_progress = None

        for ek in active_kpis:
            last_progress = (
                ProgressEntry.objects.filter(employee_kpi=ek, date__lte=now().date())
                .order_by("-date")
                .first()
            )
            if not last_progress:
                continue

            if (
                current_last_progress is None
                or last_progress.date > current_last_progress.date
            ):
                current_last_progress = last_progress
                current_kpi = ek

        # the current kpi the emloyee is working on
        if current_kpi and current_last_progress:
            data += "\n-- Current Task (Detailed) --\n"
            data += (
                f"KPI: {current_kpi.kpi.title}\n"
                f"Progress: {current_last_progress.value}%\n"
                f"Last Update: {current_last_progress.date}\n"
                f"Notes: {current_last_progress.note}\n\n"
            )

        # Add all other active KPIs to "previous" section
        for ek in active_kpis:
            if current_kpi and ek.id != current_kpi.id:
                previous_kpis.append(ek)

        # if previous tasks are completed then dont give detail insights of it.
        if previous_kpis:
            data += "\n-- Previous Tasks (Summary Only) --\n"
            for ok in previous_kpis:
                data += f"KPI: {ok.kpi.title} — Completed or previously worked on.\n"

        data += "\n----------------------------\n"

    ai_output = generate_kpi_insights(data, mode="manager")
    ai_output_html = markdown.markdown(ai_output)

    return render(request, "ai/kpi_insights.html", {"ai_output_html": ai_output_html})


## AI feature for admin to get insights of Manager, Department and Employee like a detail insight of each thing.
@login_required
@role_required(["admin"])
def ai_admin_insights(request):

    data = "=== ORGANIZATION-WIDE KPI REPORT ===\n"

    data += "\n## DEPARTMENT OVERVIEW\n"

    for dep_code, dep_name in DEPARTMENT:
        employees = EmployeeProfile.objects.filter(department=dep_code, role="employee")

        total_employees = employees.count()
        emp_kpis = EmployeeKpi.objects.filter(employee__in=employees)

        total_kpis = emp_kpis.count()
        active_kpis = [ek for ek in emp_kpis if ek.status() != "Complete"]
        completed_kpis = [ek for ek in emp_kpis if ek.status() == "Complete"]

        # average progress
        progress_values = [ek.progress_percentage() for ek in active_kpis]
        avg_progress = (
            sum(progress_values) / len(progress_values) if progress_values else 0
        )

        # risky KPIs (low progress)
        risky_kpis = [ek for ek in active_kpis if ek.progress_percentage() < 40]

        data += f"""
### {dep_name}
Employees: {total_employees}
Total KPIs: {total_kpis}
Active KPIs: {len(active_kpis)}
Completed KPIs: {len(completed_kpis)}
Average Progress: {avg_progress:.1f}%
At-Risk KPIs: {len(risky_kpis)}
"""
    ##manager summary
    data += "\n## MANAGER PERFORMANCE SUMMARY\n"

    managers = EmployeeProfile.objects.filter(role="manager")

    for mgr in managers:
        team = EmployeeProfile.objects.filter(manager=mgr.user)

        total_team = team.count()
        team_kpis = EmployeeKpi.objects.filter(employee__in=team)

        active_kpis = [ek for ek in team_kpis if ek.status() != "Complete"]
        progress_values = [ek.progress_percentage() for ek in active_kpis]

        avg_progress = (
            sum(progress_values) / len(progress_values) if progress_values else 0
        )

        data += f"""
### Manager: {mgr.user.get_full_name()}
Team Size: {total_team}
Active KPIs: {len(active_kpis)}
Average Team Progress: {avg_progress:.1f}%
"""

    data += "\n## EMPLOYEE CURRENT TASK SNAPSHOT\n"

    all_employees = EmployeeProfile.objects.filter(role="employee")

    for emp in all_employees:
        latest_entry = (
            ProgressEntry.objects.filter(
                employee_kpi__employee=emp, date__lte=now().date()
            )
            .order_by("-date")
            .first()
        )

        if latest_entry:
            current_kpi = latest_entry.employee_kpi
            data += (
                f"\nEmployee: {emp.user.get_full_name()}\n"
                f"Current KPI: {current_kpi.kpi.title}\n"
                f"Progress: {latest_entry.value}%\n"
                f"Last Update: {latest_entry.date}\n"
                f"Notes: {latest_entry.note}\n"
            )
        else:
            data += f"\nEmployee: {emp.user.get_full_name()}\nNo active KPI.\n"

    ai_output = generate_kpi_insights(data, mode="admin")
    ai_output_html = markdown.markdown(ai_output)

    return render(request, "ai/admin_insights.html", {"ai_output_html": ai_output_html})


# ai feature for employee
@login_required
@role_required(["employee"])
def ai_employee_coach(request):
    employee = EmployeeProfile.objects.get(user=request.user)
    assignments = EmployeeKpi.objects.filter(employee=employee)

    data = f"=== KPI DATA FOR {employee.user.get_full_name()} ===\n"

    for ek in assignments:
        # the recent progress
        last = (
            ProgressEntry.objects.filter(employee_kpi=ek, date__lte=now().date())
            .order_by("-date")
            .first()
        )

        progress = ek.progress_percentage()
        deadline = ek.end_date
        status = ek.status()

        data += f"""
KPI: {ek.kpi.title}
Progress: {progress}%
Status: {status}
Deadline: {deadline}
Last Update: {last.date if last else 'No updates yet'}
Notes: {last.note if last else 'No notes available'}
"""

    ai_output = generate_kpi_insights(data, mode="employee")
    ai_output_html = markdown.markdown(ai_output)

    return render(request, "ai/employee_ai.html", {"ai_output_html": ai_output_html})
